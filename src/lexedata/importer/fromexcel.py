# -*- coding: utf-8 -*-
import openpyxl as op
import csv
from pathlib import Path

from lexedata.importer.objects import Language, Concept, Form, CogSet, CognateJudgement
from lexedata.importer.database import create_db_session
from lexedata.importer.cellparser import CellParser
import lexedata.importer.exceptions as ex

create_db_session()

class ExcelParser:
    def language_from_column(self, column):
        excel_name, curator = [cell.value or "" for cell in column]
        name, comment = "", ""
        return {
            "name": name,
            "curator": curator,
            "comment": comment,
            "raw name": excel_name,
        }

    def concept_from_row(self, row):
        set, english, english_strict, spanish, portuguese, french = [cell.value or "" for cell in row]
        return {
            "set": set,
            "english": english,
            "english_strict": english_strict,
            "spanish": portuguese,
            "french": french,
            "gloss": english,
        }

    def init_lan(self, lan_iter):
        for lan_col in lan_iter:
            # iterate over language columns
            language_properties = self.language_from_column(lan_col)
            raw_name = language_properties.pop("raw name")
            id = Language.register_new_id(raw_name)
            self.lan_dict[raw_name] = Language(ID=id, **language_properties)

    def get_cell_comment(self, cell):
        return cell.comment.content if cell.comment else ""

    def init_con_form(self, con_iter, form_iter, wb):
        with (self.path / "form_init.csv").open( "w", encoding="utf8", newline="") as formsout, \
                (self.path / "concept_init.csv").open( "w", encoding="utf8", newline="") as conceptsout:

            for row_forms, row_con in zip(form_iter, con_iter):
                concept_properties = self.concept_from_row(row_con)
                concept_id = Concept.register_new_id(concept_properties.pop("gloss"))
                comment = self.get_cell_comment(row_con[0])
                concept = Concept(ID=concept_id, comment=comment, **concept_properties)

                for f_cell in row_forms:
                    if f_cell.value:
                        # get corresponding language_id to column
                        this_lan = self.lan_dict[wb[(f_cell.column_letter + "1")].value]

                        for f_ele in self.cell_parser.parse(f_cell):
                            form_cell = self.form_from_cell(f_ele, this_lan, f_cell)
                            # FIXME: Replace this by [look up existing form, otherwise create form]
                            form_id = Form.register_new_id("{:}_{:}".format(this_lan.ID, concept.ID))
                            form = Form(ID=form_id, **form_cell)
                            form.concepts.append(concept)

    def form_from_cell(self, f_ele, lan, form_cell):

        phonemic, phonetic, ortho, comment, source = f_ele

        # replace source if not given
        source_id = "{:s}_{:}".format(lan.ID, (source if source else "{1}").strip())

        return {
            "language": lan,
            "phonemic": phonemic,
            "phonetic": phonetic,
            "orthographic": ortho,
            "comment": comment,
            "sources": [source_id],
            #"procedural_comment": self.get_cell_comment(form_cell),
        }



    def initialize_lexical(self):

        wb = op.load_workbook(filename=self.lexicon_spreadsheet)
        sheets = wb.sheetnames
        wb = wb[sheets[0]]
        iter_forms = wb.iter_rows(min_row=3, min_col=7, max_col=44)  # iterates over rows with forms
        iter_concept = wb.iter_rows(min_row=3, max_col=6)  # iterates over rows with concepts
        iter_lan = wb.iter_cols(min_row=1, max_row=2, min_col=7, max_col=44)

        self.init_lan(iter_lan)
        self.init_con_form(iter_concept, iter_forms, wb)

    def cogset_from_row(self, cog_row):
        values = [cell.value or "" for cell in cog_row]
        return {"ID": values[1],
                "properties": values[0],
                "comment": self.get_cell_comment(cog_row[1])}

    def cogset_cognate(self, cogset_iter, cog_iter, lan_dict, wb):

        for cogset_row, row_forms in zip(cogset_iter, cog_iter):
            if not cogset_row[1].value:
                continue
            elif cogset_row[1].value.isupper():
                properties = self.cogset_from_row(cogset_row)
                cogset = CogSet(**properties)

                for f_cell in row_forms:
                    if f_cell.value:
                        # get corresponding language_id to column
                        this_lan = self.lan_dict[wb[(f_cell.column_letter + "1")].value]

                        try:
                            for f_ele in self.cell_parser.parse(f_cell):
                                form_cell = self.form_from_cell(f_ele, this_lan, f_cell)
                                # FIXME: Replace this by [look up existing form, otherwise create form]
                                form = Form(**form_cell)
                                CognateJudgement
                        except ex.CellParsingError as e:
                            print("{:s}{:d}:".format(f_cell.column_letter, f_cell.row), e)
                            continue
            else:
                continue


    def initialize_cognate(self):
        wb = op.load_workbook(filename=self.cognatesets_spreadsheet)
        try:
            for sheet in wb.sheetnames:
                print("\nParsing sheet '{:s}'".format(sheet))
                ws = wb[sheet]
                iter_cog = ws.iter_rows(min_row=3, min_col=5, max_col=42)  # iterates over rows with forms
                iter_congset = ws.iter_rows(min_row=3, max_col=4)  # iterates over rows with concepts
                self.cogset_cognate(iter_congset, iter_cog, self.lan_dict, ws)
        except KeyError:
            pass

    def __init__(self,
               output=Path("initial_data"),
               lexicon_spreadsheet = "TG_comparative_lexical_online_MASTER.xlsx",
               cognatesets_spreadsheet = "TG_cognates_online_MASTER.xlsx"):
        self.path = Path(output)
        if not self.path.exists():
            self.path.mkdir()
        self.lan_dict = {}
        self.lexicon_spreadsheet = lexicon_spreadsheet
        self.cognatesets_spreadsheet = cognatesets_spreadsheet
        self.cell_parser = CellParser()

    def parse(self):
        self.initialize_lexical()
        self.initialize_cognate()


if __name__ == "__main__":
    ExcelParser()
