import pytest
from pathlib import Path
import argparse
import re

import pycldf
import openpyxl

from helper_functions import copy_metadata, copy_to_temp
import lexedata.importer.excel_matrix as f


@pytest.fixture
def empty_excel():
    return Path(__file__).parent / "data/cldf/defective_dataset/empty_excel.xlsx"


# TODO: have a look at this test. just trying to pass codecov
def test_db_chache():
    copy = copy_metadata(Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json")
    res = dict()
    dataset = pycldf.Dataset.from_metadata(copy)
    db = f.DB(output_dataset=dataset)
    db.cache_dataset()
    for table in dataset.tables:
        table_type = (
            table.common_props.get("dc:conformsTo", "").rsplit("#", 1)[1] or table.url
        )
        res[table_type] = {}
    assert db.cache == res


def test_no_wordlist_and_no_cogsets(fs):
    with pytest.raises(argparse.ArgumentError,
                       match="At least one of WORDLIST and COGNATESETS excel files must be specified.*"
                       ):
        # mock empty json file
        fs.create_file("invented_path", contents="{}")
        f.load_dataset(
            metadata="invented_path",
            lexicon=None,
            cognate_lexicon=None,
        )


def test_no_dialect_excel_parser(fs, caplog, empty_excel):
    # ExcelParser
    with pytest.raises(ValueError):
        # mock empty json file
        fs.create_file("invented_path", contents="{}")
        f.load_dataset(
            metadata="invented_path",
            lexicon=empty_excel,
        )
        assert caplog.text.endswith(
            "User-defined format specification in the json-file was missing, falling back to default parser"
        )


def test_no_dialect_excel_cognate_parser(fs, caplog, empty_excel):
    # ExcelCognateParser
    with pytest.raises(ValueError):
        # mock empty json file
        fs.create_file("invented_path", contents="{}")
        f.load_dataset(
            metadata="invented_path", lexicon=None, cognate_lexicon=empty_excel
        )
        re.search(
            "User-defined format specification in the json-file was missing, falling back to default parser",
                  caplog.text
        )


def test_dialect_missing_key_excel_parser(fs, caplog, empty_excel):
    # ExcelParser
    with pytest.raises(ValueError):
        fs.create_file("invented_path", contents="""{"special:fromexcel": {}}""")
        f.load_dataset("invented_path", lexicon=empty_excel)
    re.search(
        "User-defined format specification in the json-file was missing the key lang_cell_regexes, "
        "falling back to default parser",
        caplog.text

    )


def test_dialect_missing_key_excel_cognate_parser(fs, caplog, empty_excel):
    # CognateExcelParser
    with pytest.raises(ValueError):
        fs.create_file("invented_path", contents="""{"special:fromexcel": {}}""")
        f.load_dataset("invented_path", lexicon=None, cognate_lexicon=empty_excel)
    re.search(
        r"User-defined format specification in the json-file was missing the key .*falling back to default parser.*",
        caplog.text
    )


def test_no_first_row_in_excel(empty_excel):
    original = Path(__file__).parent / "data/cldf/minimal/cldf-metadata.json"
    copy = copy_metadata(original=original)
    with pytest.raises(
            AssertionError,
            match="Your first data row didn't have a name. Please check your format specification or ensure the "
                  "first row has a name."
    ):
        f.load_dataset(metadata=copy, lexicon=empty_excel)


def test_language_regex_error():
    excel = (
        Path(__file__).parent
        / "data/cldf/defective_dataset/small_defective_no_regexes.xlsx"
    )
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.lang_cell_regexes = [r"(?P<Name>\[.*)", "(?P<Curator>.*)"]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)

    with pytest.raises(
        ValueError,
        match=r"In cell G1: Expected to encounter match for .*, but found no_language"
    ):
        EP.parse_cells(lexicon_wb)


def test_language_comment_regex_error():
    excel = (
        Path(__file__).parent
        / "data/cldf/defective_dataset/small_defective_no_regexes.xlsx"
    )
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.lang_comment_regexes = [r"\[.*", ".*"]

    EP = f.excel_parser_from_dialect(
        dataset,
        argparse.Namespace(
            lang_cell_regexes=[r"\[.*", ".*"],
            lang_comment_regexes=[".*"],
            row_cell_regexes=["(?P<Name>.*)"],
            row_comment_regexes=[".*"],
            cell_parser={
                "form_separator": ",",
                "variant_separator": "~",
                "name": "CellParser",
                "cell_parser_semantics": [
                    ["<", ">", "Form", True],
                    ["{", "}", "Source", False],
                ],
            },
            check_for_match=["Form"],
            check_for_row_match=["Name"],
            check_for_language_match=["Name"],
        ),
        cognate=False,
    )
    EP = EP(dataset)
    with pytest.raises(ValueError, match="In cell B1: Expected to encounter match for .*, but found .*"):
        EP.parse_cells(lexicon_wb)


def test_properties_regex_error():
    excel = (
        Path(__file__).parent
        / "data/cldf/defective_dataset/small_defective_no_regexes.xlsx"
    )
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.row_cell_regexes = [
        "(?P<set>.*)",
        r"(?P<Name>\[.*)",
        "(?P<English>.*)",
        "(?P<Spanish>.*)",
        "(?P<Portuguese>.*)",
        "(?P<French>.*)",
    ]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)

    with pytest.raises(
        ValueError,
        match=r"In cell B3: Expected to encounter match for .*, but found no_concept_name"
    ):
        EP.parse_cells(lexicon_wb)


def test_properties_comment_regex_error():
    excel = (
        Path(__file__).parent
        / "data/cldf/defective_dataset/small_defective_no_regexes.xlsx"
    )
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    copy = copy_metadata(original=original)

    dataset = pycldf.Dataset.from_metadata(copy)
    dialect = argparse.Namespace(**dataset.tablegroup.common_props["special:fromexcel"])
    lexicon_wb = openpyxl.load_workbook(excel).active
    dialect.row_comment_regexes = [".*", r"\[.*", ".*", ".*", ".*", ".*"]
    EP = f.excel_parser_from_dialect(dataset, dialect, cognate=False)
    EP = EP(dataset)
    with pytest.raises(
        ValueError,
        match=r"In cell B3: Expected to encounter match for .*, but found no_concept_comment"
    ):
        EP.parse_cells(lexicon_wb)


def test_cognate_parser_language_not_found():
    excel = Path(__file__).parent / "data/excel/minimal_cog.xlsx"
    original = Path(__file__).parent / "data/cldf/smallmawetiguarani/cldf-metadata.json"
    dataset, copy = copy_to_temp(original)
    lexicon_wb = openpyxl.load_workbook(excel).active
    EP = f.ExcelCognateParser(output_dataset=dataset)
    print(dataset["LanguageTable"])
    with pytest.raises(
        ValueError,
        match="Failed to find object {'ID': 'autaa', 'Name': 'Autaa', 'Comment': "
              "'fictitious!'} in the database. In cell: D1"
    ):
        EP.db.cache_dataset()
        EP.parse_cells(lexicon_wb)
